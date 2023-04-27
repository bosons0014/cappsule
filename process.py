#!/usr/local/bin/python3
#cappsule 
import openpyxl

path = "./data.xlsx"

obj = openpyxl.load_workbook(path)

master_sheet = obj["Master"]
test_sheet = obj["Test"]
obj.create_sheet("Output2")
output_sheet = obj["Output2"]
hash = dict([])
row = 1
#prepare a hash table with key = 
#1. Medium
#2. Quantity
#3. Manufacturer
#4. Composition

while row < master_sheet.max_row:
    if row == 1:
       row = row + 1
       continue
    hash[str(master_sheet.cell(row = row, column = 4).value).lower().replace("/","").replace("-","").replace(" ","").replace("+","").replace("and","") + str(master_sheet.cell(row = row, column = 5).value).lower().replace("/","").replace("-","").replace(" ","").replace("+","").replace("and","") + str(master_sheet.cell(row = row, column = 6).value).lower().replace("/","").replace("-","").replace(" ","").replace("+","").replace("and","")+ str(master_sheet.cell(row = row, column = 7).value).lower().replace("/","").replace("-","").replace(" ","").replace("+","").replace("and","")] = row 
    row = row + 1
output_sheet.cell(row = 1, column = 1).value = test_sheet.cell(row = 1, column = 1).value
output_sheet.cell(row = 1, column = 2).value = test_sheet.cell(row = 1, column = 2).value
output_sheet.cell(row = 1, column = 3).value = test_sheet.cell(row = 1, column = 3).value
output_sheet.cell(row = 1, column = 4).value = test_sheet.cell(row = 1, column = 4).value
output_sheet.cell(row = 1, column = 5).value = test_sheet.cell(row = 1, column = 5).value
output_sheet.cell(row = 1, column = 6).value = test_sheet.cell(row = 1, column = 6).value
output_sheet.cell(row = 1, column = 7).value = master_sheet.cell(row = 1, column = 2).value
_row2_ = 1

#based on the hash table 
#lookup the keys
#write it out to the output sheet
#in a new worksheet
#save your work
#also make sure to filter/massage the input keys for better match

while _row2_ <= test_sheet.max_row:
    if _row2_ == 1:
        _row2_ = _row2_ + 1
        continue
    else:
        key = str(test_sheet.cell(row = _row2_, column = 3).value).lower().replace("/","").replace("-","").replace(" ","").replace("+","").replace("and","") + str(test_sheet.cell(row = _row2_, column = 4).value).lower().replace("/","").replace("-","").replace(" ","").replace("+","").replace("and","") + str(test_sheet.cell(row = _row2_, column = 5).value).lower().replace("/","").replace("-","").replace(" ","").replace("+","").replace("and","") + str(test_sheet.cell(row = _row2_, column = 6).value).lower().replace("/","").replace("-","").replace(" ","").replace("+","").replace("and","")      
        if key in hash.keys():
              output_sheet.cell(row = _row2_, column = 1).value = test_sheet.cell(row = _row2_, column = 1).value
              output_sheet.cell(row = _row2_, column = 2).value = test_sheet.cell(row = _row2_, column = 2).value
              output_sheet.cell(row = _row2_, column = 3).value = test_sheet.cell(row = _row2_, column = 3).value
              output_sheet.cell(row = _row2_, column = 4).value = test_sheet.cell(row = _row2_, column = 4).value
              output_sheet.cell(row = _row2_, column = 5).value = test_sheet.cell(row = _row2_, column = 5).value
              output_sheet.cell(row = _row2_, column = 6).value = test_sheet.cell(row = _row2_, column = 6).value
              output_sheet.cell(row = _row2_, column = 7).value = master_sheet.cell(row = hash[key], column = 2).value
    _row2_ = _row2_ + 1
obj.save("data.xlsx")
